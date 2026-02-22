import sys
import json
import re
import os
import datetime
import shutil
import openpyxl
from openpyxl.styles import Alignment

# --- SABİTLER ---
TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), 'templates')
EK2_TEMPLATE = os.path.join(TEMPLATE_DIR, 'EK-2.xlsx')
EK3_TEMPLATE = os.path.join(TEMPLATE_DIR, 'EK-3.xlsx')

def get_desktop_path(custom_path=None):
    if custom_path:
        return custom_path
    return os.path.join(os.environ['USERPROFILE'], 'Desktop')

def get_sort_key(text):
    try:
        # "1.1 - Konu" -> "1.1" -> "1.10" -> [1, 10]
        code_part = text.split('-')[0].strip()
        parts_str = code_part.split('.')
        parts = []
        for i, p in enumerate(parts_str):
            if i == 1 and len(p) == 1: 
                 p = p + "0"
            parts.append(int(p))
        return parts
    except:
        return [9999, 9999]

def run_process(data_json):
    try:
        if isinstance(data_json, bytes):
            data_json = data_json.decode('utf-8')
        
        payload = json.loads(data_json)
        islem_tipi = payload.get('tip') # 'EK2' veya 'EK3'
        veri = payload.get('veri', {})
        desktop_path = payload.get('desktop_path')
        
        target_template = EK2_TEMPLATE if islem_tipi == 'EK2' else EK3_TEMPLATE
        
        if not os.path.exists(target_template):
            return {"success": False, "error": f"Şablon dosyası bulunamadı: {target_template}"}

        # Çıktı dosyasını oluştur
        timestamp = datetime.datetime.now().strftime('%d%m%Y_%H%M%S')
        output_filename = f"{islem_tipi}_{timestamp}.xlsx"
        output_path = os.path.join(get_desktop_path(desktop_path), output_filename)
        
        # Şablonu kopyala
        shutil.copy(target_template, output_path)
        
        # Dosyayı aç
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active # Varsayılan olarak ilk sayfayı al

        if islem_tipi == 'EK2':
            program_adi = veri.get('program_adi', '')
            dersler = veri.get('dersler', [])
            
            # Veri Hazırlığı
            dersler_sorted = sorted(dersler, key=lambda x: (x.get('tarih'), x.get('saat')))
            dates = sorted(list(set([d.get('tarih') for d in dersler])))
            
            # Tarihleri formatla
            # tr_dates = [datetime.datetime.strptime(d, '%Y-%m-%d').strftime('%d.%m.%Y') for d in dates]
            
            unique_pairs = []
            seen = set()
            for l in dersler_sorted:
                pair = (l.get('konu'), l.get('egitici'))
                if pair not in seen:
                    seen.add(pair)
                    unique_pairs.append(pair)
            
            unique_pairs.sort(key=lambda x: get_sort_key(x[0]))
            
            # Hücrelere Yaz
            ws['A5'] = f"Eğitimin Adı: {program_adi} "
            ws['A6'] = f"Tarihi: {datetime.datetime.now().strftime('%d.%m.%Y')} "

            # Temizle: B13'ten B52'ye kadar (dahil)
            for row_num in range(13, 53):
                ws.cell(row=row_num, column=2).value = None
            
            # Liste Yazımı (B13'ten başla)
            start_row = 13
            for i, (subj, trainer_raw) in enumerate(unique_pairs):
                clean_trainer = trainer_raw.split(" - ")[0] if " - " in trainer_raw else trainer_raw
                
                # Konu
                cell_subj = ws.cell(row=start_row + (i * 2), column=2) # B sütunu
                cell_subj.value = subj
                
                # Eğitici (Bir alt satıra)
                cell_trainer = ws.cell(row=start_row + (i * 2) + 1, column=2) # B sütunu
                cell_trainer.value = clean_trainer
                
        elif islem_tipi == 'EK3':
            program_adi = veri.get('program_adi', '')
            dersler = veri.get('dersler', [])
            personeller = veri.get('personeller', [])
            duzenleyen = veri.get('duzenleyen', {})
            onaylayan = veri.get('onaylayan', {})
            
            dates = sorted(list(set([d.get('tarih') for d in dersler])))
            subjects = list(set([d.get('konu') for d in dersler]))
            codes = []
            for s in subjects:
                # 1. " - " ile ayrılmışsa, ilk kısmı al
                if ' - ' in s:
                    code = s.split(' - ')[0].strip()
                    codes.append(code)
                else:
                    # 2. Ayrılmamışsa, baştaki sayı/noktalı kısmı regex ile al
                    match = re.match(r'^([\d\.]+)', s)
                    if match:
                        code = match.group(1).rstrip('.')  # '1.2.' ise '1.2' yap
                        codes.append(code)
                    else:
                        # 3. Fallback: Sadece ilk kelimeyi al
                        codes.append(s.split()[0])
            
            # Benzersiz yap ve sırala
            codes = sorted(list(set(codes)), key=lambda x: get_sort_key(x))
            
            # Saat Hesaplama
            total_hours = 0.0
            for d in dersler:
                saat_str = d.get('saat', '')
                try:
                    if '-' in saat_str:
                        start_str, end_str = saat_str.split('-')
                        sh, sm = map(int, start_str.strip().split(':'))
                        eh, em = map(int, end_str.strip().split(':'))
                        start_min = sh * 60 + sm
                        end_min = eh * 60 + em
                        diff = end_min - start_min
                        if diff > 0:
                            total_hours += diff / 60.0
                        else:
                            total_hours += 3.0
                    else:
                        total_hours += 3.0
                except:
                    total_hours += 3.0
            
            # Cinsiyet
            kadin = 0
            erkek = 0
            for p in personeller:
                cns = p.get('cinsiyet', '').lower().strip()
                if cns in ['kadın', 'kadin', 'female', 'k']:
                    kadin += 1
                else: 
                    # Varsayılan erkek, ama temizlenmiş veri
                    erkek += 1
            toplam = kadin + erkek
            
            # Tarih Aralığı
            tr_dates = [datetime.datetime.strptime(d, '%Y-%m-%d').strftime('%d.%m.%Y') for d in dates]
            tarih_araligi = f"{tr_dates[0]} - {tr_dates[-1]}" if dates else "-"
            
            # Eğitici Dizgisi
            egitici_str = ""
            if dersler:
                raw = dersler[0].get('egitici', '')
                parts = raw.split(' - ')
                if len(parts) > 1:
                     egitici_str = f"{parts[0]}\n{parts[1]}"
                else:
                     egitici_str = raw
            
            # Hücrelere Yaz
            ws['B5'] = str(datetime.datetime.now().year)
            ws['B6'] = program_adi
            ws['B11'] = tarih_araligi
            ws['C11'] = ", ".join(codes)
            ws['D11'] = str(len(codes))
            ws['E11'] = total_hours
            
            # Katılan Sayısı (Row 11)
            ws['F11'] = kadin
            ws['G11'] = erkek
            ws['H11'] = toplam
            ws['I11'] = toplam # Başarılı
            ws['J11'] = "0"    # Başarısız
            
            # Alt Kısım - Yetkililer
            today = datetime.datetime.now().strftime("%d.%m.%Y")
            
            # Düzenleyen (A20 Ad, A21 Unvan)
            if duzenleyen:
                ws['A20'] = duzenleyen.get('ad_soyad', '')
                ws['A21'] = duzenleyen.get('unvan', '')
            
            # Tarih (D20)
            ws['D20'] = today
            
            # Eğitici (E20 Ad, E21 Unvan)
            if dersler:
                raw = dersler[0].get('egitici', '')
                parts = raw.split(' - ')
                if len(parts) > 1:
                    ws['E20'] = parts[0]
                    ws['E21'] = parts[1]
                else:
                    ws['E20'] = raw
                    ws['E21'] = ''

            # Onaylayan (H20 Ad, H21 Unvan)
            if onaylayan:
                ws['H20'] = onaylayan.get('ad_soyad', '')
                ws['H21'] = onaylayan.get('unvan', '')

        # Kaydet
        wb.save(output_path)
        return {"success": True, "path": output_path}

    except Exception as e:
        import traceback
        return {"success": False, "error": f"{type(e).__name__}: {str(e)}\n{traceback.format_exc()}"}

if __name__ == "__main__":
    try: sys.stdin.reconfigure(encoding='utf-8')
    except: pass
    
    try:
        input_data = sys.stdin.read()
        if input_data:
            result = run_process(input_data)
            print(json.dumps(result, ensure_ascii=False))
        else:
            print(json.dumps({"success": False, "error": "Girdi verisi boş"}, ensure_ascii=False))
    except Exception as e:
        print(json.dumps({"success": False, "error": str(e)}, ensure_ascii=False))
